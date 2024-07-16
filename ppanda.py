import pandas as pd
from openpyxl import load_workbook

# Carregar o arquivo Excel
arquivo_excel = "caminho_para_seu_arquivo.xlsx"
wb = load_workbook(arquivo_excel)

# Escolher a planilha (sheet) desejada, supondo que seja a primeira
sheet_name = wb.sheetnames[0]

# Carregar a planilha como um DataFrame do pandas
df = pd.read_excel(arquivo_excel, sheet_name=sheet_name)

# Supondo que a coluna que você quer verificar se chama "Resposta"
coluna = "Resposta"

# Contar quantas vezes "concordam" aparece na coluna
contagem_concordam = (df[coluna] == "concordam").sum()

print(f"Número de 'concordam' na coluna '{coluna}': {contagem_concordam}")